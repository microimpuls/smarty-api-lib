#!/usr/bin/python3
# -*- coding: utf-8 -*-
import csv
# for parsing xlsx
import openpyxl # pip install openpyxl
try:
    from client import SmartyContentAPI
except ImportError:
    raise ImportError("client.py file not found")

SERVER_HOST = 'http://127.0.0.1:8000'
API_KEY = 'api_key' # Content API key
CLIENT_ID = 1
INPUT_FILE_NAME = 'path to the file'

class ImportVideoFromFile():
    def __init__(self, api, file_path: str):
        self.api = api
        self.file_path = file_path

        if self.file_path.split('.')[-1] == "csv":
            self.parse_csv()
        elif self.file_path.split('.')[-1] == "xlsx":
            self.parse_xlsx()
        else:
            print("Exctption: type file not a .csv or .xlsx")

    def parse_csv(self):
        kwargs = {}
        with open(self.file_path, 'r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                name = row.pop('name')
                rating = row.pop('rating')
                if not name or not rating:
                    print("Exception: not name or rating")
                    continue
                for key, value in row.items():
                    kwargs[key] = value
                self.api.video_create(name, rating, **kwargs)
                kwargs.clear()

    def parse_xlsx(self):
        wookbook = openpyxl.load_workbook(self.file_path)
        worksheet = wookbook.active
        kwargs = {}
        for i in range(1, worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column):
                if col[i].value and col[0].value:
                    kwargs[col[0].value] = col[i].value
            name = kwargs.pop('name')
            rating = kwargs.pop('rating')
            self.api.video_create(name, rating, **kwargs)
            kwargs.clear()

if __name__ == "__main__":
    api = SmartyContentAPI(SERVER_HOST, CLIENT_ID, API_KEY)
    ImportVideoFromFile(api, INPUT_FILE_NAME)